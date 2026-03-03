#!/usr/bin/env python3
"""
export_annual_pnl.py
====================

Purpose
-------
Union multiple bet-history sources into a normalized dataset and export an
accountant-friendly Excel workbook summarizing annual PnL.

Required Input
--------------
- Pikkit CSV export is the ONLY required input.
  The script will error if:
    - inputs.pikkit is missing/null
    - inputs.pikkit.enabled is not True
    - inputs.pikkit.csv_path is missing/empty

Optional Inputs
---------------
- OddsJam Bet Tracker CSV (inputs.oddsjam can be null)
- Manual tracker CSV (inputs.manual can be null)

Key Features
------------
- Tax-year filtering (per source)
- Optional OddsJam sportsbook whitelist filtering (optionally date-range scoped)
- Profit computation when profit is not provided (OddsJam/manual)
- Excel workbook output:
    - Summary (overall)
    - By Month (optional)
    - Grouped rollups (configurable)
    - Raw Union sheet (optional)

Important Excel Note
--------------------
Excel does not support timezone-aware datetimes. We parse timestamps with utc=True
for correctness, then convert to timezone-naive UTC before writing to Excel.

Run
---
poetry run python scripts/export_annual_pnl.py --config configs/annual_pnl_export.yaml
"""

from __future__ import annotations

import argparse
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import yaml  # pyyaml
except ImportError as e:
    raise ImportError(
        "Missing dependency: pyyaml. Install with `poetry add pyyaml` or `pip install pyyaml`."
    ) from e

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# -----------------------------
# Normalized union schema
# -----------------------------
NORMALIZED_COLUMNS = [
    "source",  # pikkit | oddsjam | manual
    "bet_id",  # best-effort (pikkit has it; oddsjam/manual may be NA)
    "created_dt",  # UTC timestamp (timezone-naive before Excel export)
    "created_date",  # YYYY-MM-DD (string)
    "tax_month",  # YYYY-MM (string)
    "sportsbook",
    "sport",
    "league",
    "market_name",
    "bet_name",
    "bet_type",
    "status",  # WIN | LOSS | PUSH | VOID | OTHER
    "odds",  # American odds numeric
    "stake",  # numeric
    "potential_payout",  # stake * decimal_odds (if odds available)
    "bet_profit",  # realized profit
]


@dataclass(frozen=True)
class SourceCfg:
    """
    Per-source configuration.

    Note: oddsjam/manual may be omitted or set to null in YAML. In that case, they are skipped.
    """

    enabled: bool
    csv_path: str

    # OddsJam-only optional config
    sportsbooks_include: list[str] | None = None
    sportsbooks_include_mode: str | None = None
    sportsbooks_include_date_range: dict[str, str] | None = None
    column_map: dict[str, list[str]] | None = None
    status_map: dict[str, list[str]] | None = None


@dataclass(frozen=True)
class Config:
    """
    Top-level configuration loaded from YAML.
    """

    output_xlsx: str
    tax_year: int
    tax_year_date_field_by_source: dict[str, str]
    inputs: dict[str, SourceCfg | None]  # oddsjam/manual can be None
    include_raw_union_sheet: bool
    include_monthly_sheet: bool
    groupings: list[list[str]]
    round_decimals: int


# -----------------------------
# Helpers
# -----------------------------
def _first_present_col(df: pd.DataFrame, candidates: Sequence[str]) -> str | None:
    cols = set(df.columns)
    for c in candidates:
        if c in cols:
            return c
    return None


def _to_float(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _make_excel_safe_datetime(df: pd.DataFrame, col: str = "created_dt") -> pd.DataFrame:
    """
    Excel does not support timezone-aware datetimes.

    If df[col] is tz-aware, convert to timezone-naive while preserving UTC clock time:
      tz-aware UTC -> tz-naive (same timestamps displayed, but tz removed)
    """
    if col not in df.columns:
        return df

    if not pd.api.types.is_datetime64_any_dtype(df[col]):
        return df

    # If tz-aware, drop tz
    try:
        if getattr(df[col].dt, "tz", None) is not None:
            df[col] = df[col].dt.tz_convert("UTC").dt.tz_localize(None)
    except Exception:
        # best-effort fallback
        s = pd.to_datetime(df[col], utc=True, errors="coerce")
        df[col] = s.dt.tz_convert("UTC").dt.tz_localize(None)

    return df


def american_to_decimal(odds: float) -> float | None:
    """
    Convert American odds to decimal odds. Returns None if invalid.
    """
    if odds is None or pd.isna(odds):
        return None
    try:
        o = float(odds)
    except Exception:
        return None
    if o == 0:
        return None
    if o > 0:
        return 1.0 + (o / 100.0)
    return 1.0 + (100.0 / abs(o))


def normalize_status(raw: Any, status_map: dict[str, list[str]]) -> str:
    """
    Normalize status values into WIN/LOSS/PUSH/VOID/OTHER using a config-driven map.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return "OTHER"
    s = str(raw).strip().upper()

    def _match(key: str) -> bool:
        return any(s == str(v).strip().upper() for v in status_map.get(key, []))

    if _match("win"):
        return "WIN"
    if _match("loss"):
        return "LOSS"
    if _match("push"):
        return "PUSH"
    if _match("void"):
        return "VOID"
    return "OTHER"


def compute_profit_and_payout_from_status_odds_stake(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compute:
      - potential_payout = stake * decimal_odds
      - bet_profit based on status:
          WIN  -> stake*(decimal_odds - 1)
          LOSS -> -stake
          PUSH/VOID -> 0
          OTHER -> NA
    """
    odds_dec = df["odds"].apply(american_to_decimal)
    df["potential_payout"] = df["stake"] * odds_dec

    win_profit = df["stake"] * (odds_dec - 1.0)
    df["bet_profit"] = df.get("bet_profit", pd.Series([pd.NA] * len(df)))
    df.loc[df["status"] == "WIN", "bet_profit"] = win_profit[df["status"] == "WIN"]
    df.loc[df["status"] == "LOSS", "bet_profit"] = -df.loc[df["status"] == "LOSS", "stake"]
    df.loc[df["status"].isin(["PUSH", "VOID"]), "bet_profit"] = 0.0
    return df


def _parse_date_yyyymmdd(s: str | None) -> pd.Timestamp | None:
    """
    Parse YYYY-MM-DD into UTC midnight Timestamp. Returns None if invalid.
    """
    if not s:
        return None
    ts = pd.to_datetime(s, errors="coerce")
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts.date(), tz="UTC")


def apply_oddsjam_sportsbook_include_filter(
    df: pd.DataFrame,
    sportsbook_col: str,
    created_dt: pd.Series,
    sportsbooks_include: list[str] | None,
    mode: str,
    date_range: dict[str, str] | None,
) -> pd.DataFrame:
    """
    Apply OddsJam sportsbook include filtering with optional date-range semantics.

    Modes:
      - whitelist_global:
          Always keep only rows where sportsbook is in sportsbooks_include
      - whitelist_only:
          Within date range keep only include list; outside date range keep all
      - whitelist_date_range_only:
          Keep only rows where sportsbook in include list AND created_dt in date range

    If sportsbooks_include is empty/None, df is returned unchanged.
    If date range is required by mode but invalid/missing, falls back to whitelist_global.
    """
    if not sportsbooks_include:
        return df

    include_norm = {s.strip().lower() for s in sportsbooks_include}
    sb = df[sportsbook_col].astype(str).str.strip().str.lower()
    is_in_list = sb.isin(include_norm)

    mode = (mode or "whitelist_global").strip().lower()
    if mode == "whitelist_global":
        return df[is_in_list].copy()

    start = _parse_date_yyyymmdd((date_range or {}).get("start_date"))
    end = _parse_date_yyyymmdd((date_range or {}).get("end_date"))
    if start is None or end is None:
        return df[is_in_list].copy()

    end_excl = end + pd.Timedelta(days=1)
    in_range = (created_dt >= start) & (created_dt < end_excl)

    if mode == "whitelist_only":
        keep = (~in_range) | (in_range & is_in_list)
        return df[keep].copy()

    if mode == "whitelist_date_range_only":
        keep = in_range & is_in_list
        return df[keep].copy()

    return df[is_in_list].copy()


# -----------------------------
# Config loading
# -----------------------------
def _parse_source_cfg(src: Any) -> SourceCfg | None:
    """
    Parse a per-source config node. Returns None if src is null.

    Important:
    - This function does NOT enforce that the source is required.
      The requirement (Pikkit must be enabled) is enforced in main().
    """
    if src is None:
        return None
    enabled = bool(src.get("enabled", False))
    csv_path = str(src.get("csv_path", "")).strip()
    # If enabled but no path, fail fast.
    if enabled and not csv_path:
        raise ValueError("A source is enabled but csv_path is missing/empty.")
    return SourceCfg(
        enabled=enabled,
        csv_path=csv_path,
        sportsbooks_include=src.get("sportsbooks_include"),
        sportsbooks_include_mode=src.get("sportsbooks_include_mode"),
        sportsbooks_include_date_range=src.get("sportsbooks_include_date_range"),
        column_map=src.get("column_map"),
        status_map=src.get("status_map"),
    )


def load_config(path: str) -> Config:
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f)

    required_top = ["output_xlsx", "tax_year", "tax_year_date_field_by_source", "inputs"]
    for k in required_top:
        if k not in data:
            raise ValueError(f"Config missing required key: {k}")

    inputs_raw = data["inputs"]
    inputs: dict[str, SourceCfg | None] = {
        "pikkit": _parse_source_cfg(inputs_raw.get("pikkit")),
        "oddsjam": _parse_source_cfg(inputs_raw.get("oddsjam")),
        "manual": _parse_source_cfg(inputs_raw.get("manual")),
    }

    groupings = data.get("groupings", [["sportsbook"]])
    if not isinstance(groupings, list) or not all(isinstance(g, list) for g in groupings):
        raise ValueError("Config key `groupings` must be a list of lists.")

    return Config(
        output_xlsx=str(data["output_xlsx"]),
        tax_year=int(data["tax_year"]),
        tax_year_date_field_by_source=dict(data["tax_year_date_field_by_source"]),
        inputs=inputs,
        include_raw_union_sheet=bool(data.get("include_raw_union_sheet", True)),
        include_monthly_sheet=bool(data.get("include_monthly_sheet", True)),
        groupings=groupings,
        round_decimals=int(data.get("round_decimals", 2)),
    )


# -----------------------------
# Readers
# -----------------------------
def read_pikkit(
    cfg: SourceCfg, date_field: str, tax_year: int
) -> tuple[pd.DataFrame, dict[str, int]]:
    df = pd.read_csv(cfg.csv_path, low_memory=False)

    required = {
        "bet_id",
        "sportsbook",
        "type",
        "status",
        "odds",
        "amount",
        "profit",
        "time_placed_iso",
        "time_settled_iso",
        "sports",
        "leagues",
    }
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Pikkit CSV missing required columns: {sorted(missing)}")

    pikkit_status_map = {
        "win": ["SETTLED_WIN"],
        "loss": ["SETTLED_LOSS"],
        "push": ["SETTLED_PUSH"],
        "void": ["SETTLED_VOID"],
    }

    dt_col = "time_settled_iso" if date_field == "time_settled_iso" else "time_placed_iso"
    created_dt = pd.to_datetime(df[dt_col], utc=True, errors="coerce")

    out = pd.DataFrame(
        {
            "source": "pikkit",
            "bet_id": df["bet_id"].astype(str),
            "created_dt": created_dt,
            "sportsbook": df["sportsbook"],
            "sport": df["sports"],
            "league": df["leagues"],
            "market_name": pd.NA,
            "bet_name": df.get("bet_info", pd.NA),
            "bet_type": df["type"],
            "status": df["status"].apply(lambda x: normalize_status(x, pikkit_status_map)),
            "odds": _to_float(df["odds"]),
            "stake": _to_float(df["amount"]),
            "bet_profit": _to_float(df["profit"]),
        }
    )

    odds_dec = out["odds"].apply(american_to_decimal)
    out["potential_payout"] = out["stake"] * odds_dec

    before = len(out)
    out = out[out["created_dt"].dt.year == tax_year].copy()
    dropped = before - len(out)

    # Make Excel-safe BEFORE generating tax_month to avoid tz warnings
    out = _make_excel_safe_datetime(out, "created_dt")
    out["created_date"] = out["created_dt"].dt.date.astype(str)
    out["tax_month"] = out["created_dt"].dt.to_period("M").astype(str)

    out = out.reindex(columns=NORMALIZED_COLUMNS)
    return out, {
        "rows_read": before,
        "rows_kept_tax_year": len(out),
        "rows_dropped_bad_or_other_year": dropped,
    }


def read_oddsjam(cfg: SourceCfg, tax_year: int) -> tuple[pd.DataFrame, dict[str, int]]:
    df = pd.read_csv(cfg.csv_path, low_memory=False)

    if not cfg.column_map:
        raise ValueError("OddsJam enabled but inputs.oddsjam.column_map is missing in YAML.")

    cm = cfg.column_map
    created_col = _first_present_col(df, cm.get("created_date_candidates", []))
    sportsbook_col = _first_present_col(df, cm.get("sportsbook", []))
    sport_col = _first_present_col(df, cm.get("sport", []))
    league_col = _first_present_col(df, cm.get("league", []))
    market_col = _first_present_col(df, cm.get("market_name", []))
    bet_name_col = _first_present_col(df, cm.get("bet_name", []))
    odds_col = _first_present_col(df, cm.get("odds", []))
    stake_col = _first_present_col(df, cm.get("stake", []))
    status_col = _first_present_col(df, cm.get("status", []))
    bet_type_col = _first_present_col(df, cm.get("bet_type", []))
    profit_col = _first_present_col(df, cm.get("profit", []))

    missing = [
        k
        for k, v in {
            "created_date": created_col,
            "sportsbook": sportsbook_col,
            "odds": odds_col,
            "stake": stake_col,
            "status": status_col,
        }.items()
        if v is None
    ]
    if missing:
        raise ValueError(
            f"OddsJam CSV missing required detected columns: {missing}. Update column_map in YAML."
        )

    before_read = len(df)
    created_dt_all = pd.to_datetime(df[created_col], utc=True, errors="coerce")

    df_f = apply_oddsjam_sportsbook_include_filter(
        df=df,
        sportsbook_col=sportsbook_col,
        created_dt=created_dt_all,
        sportsbooks_include=cfg.sportsbooks_include,
        mode=cfg.sportsbooks_include_mode or "whitelist_global",
        date_range=cfg.sportsbooks_include_date_range,
    )
    after_sb_filter = len(df_f)

    created_dt = pd.to_datetime(df_f[created_col], utc=True, errors="coerce")

    status_map = cfg.status_map or {}
    out = pd.DataFrame(
        {
            "source": "oddsjam",
            "bet_id": pd.NA,
            "created_dt": created_dt,
            "sportsbook": df_f[sportsbook_col],
            "sport": df_f[sport_col] if sport_col else pd.NA,
            "league": df_f[league_col] if league_col else pd.NA,
            "market_name": df_f[market_col] if market_col else pd.NA,
            "bet_name": df_f[bet_name_col] if bet_name_col else pd.NA,
            "bet_type": df_f[bet_type_col] if bet_type_col else pd.NA,
            "status": df_f[status_col].apply(lambda x: normalize_status(x, status_map)),
            "odds": _to_float(df_f[odds_col]),
            "stake": _to_float(df_f[stake_col]),
            "bet_profit": _to_float(df_f[profit_col]) if profit_col else pd.NA,
        }
    )

    out = compute_profit_and_payout_from_status_odds_stake(out)

    before_year = len(out)
    out = out[out["created_dt"].dt.year == tax_year].copy()
    dropped_year = before_year - len(out)

    out = _make_excel_safe_datetime(out, "created_dt")
    out["created_date"] = out["created_dt"].dt.date.astype(str)
    out["tax_month"] = out["created_dt"].dt.to_period("M").astype(str)

    out = out.reindex(columns=NORMALIZED_COLUMNS)

    stats = {
        "rows_read": before_read,
        "rows_after_sportsbook_include_filter": after_sb_filter,
        "rows_kept_tax_year": len(out),
        "rows_dropped_bad_or_other_year": dropped_year,
    }
    return out, stats


def read_manual(cfg: SourceCfg, tax_year: int) -> tuple[pd.DataFrame, dict[str, int]]:
    df = pd.read_csv(cfg.csv_path, low_memory=False)
    required = {
        "created_date",
        "sportsbook",
        "sport",
        "league",
        "market_name",
        "bet_name",
        "odds",
        "stake",
        "status",
        "bet_type",
    }
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Manual CSV missing required columns: {sorted(missing)}")

    status_map = cfg.status_map or {}

    out = pd.DataFrame(
        {
            "source": "manual",
            "bet_id": pd.NA,
            "created_dt": pd.to_datetime(df["created_date"], utc=True, errors="coerce"),
            "sportsbook": df["sportsbook"],
            "sport": df["sport"],
            "league": df["league"],
            "market_name": df["market_name"],
            "bet_name": df["bet_name"],
            "bet_type": df["bet_type"],
            "status": df["status"].apply(lambda x: normalize_status(x, status_map)),
            "odds": _to_float(df["odds"]),
            "stake": _to_float(df["stake"]),
            "bet_profit": pd.NA,
        }
    )

    out = compute_profit_and_payout_from_status_odds_stake(out)

    before = len(out)
    out = out[out["created_dt"].dt.year == tax_year].copy()
    dropped = before - len(out)

    out = _make_excel_safe_datetime(out, "created_dt")
    out["created_date"] = out["created_dt"].dt.date.astype(str)
    out["tax_month"] = out["created_dt"].dt.to_period("M").astype(str)

    out = out.reindex(columns=NORMALIZED_COLUMNS)
    return out, {
        "rows_read": before,
        "rows_kept_tax_year": len(out),
        "rows_dropped_bad_or_other_year": dropped,
    }


# -----------------------------
# Rollups + Excel output
# -----------------------------
def build_rollup(df: pd.DataFrame, group_cols: Sequence[str]) -> pd.DataFrame:
    tmp = df.copy()
    tmp["gross_win_profit"] = tmp["bet_profit"].where(tmp["bet_profit"] > 0, 0.0)
    tmp["gross_loss_amount"] = -tmp["bet_profit"].where(tmp["bet_profit"] < 0, 0.0)

    agg = (
        tmp.groupby(list(group_cols), dropna=False)
        .agg(
            bets=("bet_profit", "count"),
            handle=("stake", "sum"),
            net_profit=("bet_profit", "sum"),
            gross_win_profit=("gross_win_profit", "sum"),
            gross_loss_amount=("gross_loss_amount", "sum"),
        )
        .reset_index()
    )
    agg["roi"] = agg["net_profit"] / agg["handle"].where(agg["handle"] != 0, pd.NA)
    agg["avg_bet"] = agg["handle"] / agg["bets"].where(agg["bets"] != 0, pd.NA)
    return agg.sort_values("net_profit", ascending=False)


def autosize_columns(ws: Worksheet, max_width: int = 60) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        values = [str(c.value) if c.value is not None else "" for c in col_cells[:2000]]
        width = min(max((len(v) for v in values), default=10) + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def apply_number_formats(
    ws: Worksheet, currency_cols: Sequence[str], percent_cols: Sequence[str]
) -> None:
    header = [c.value for c in ws[1]]
    col_map = {name: idx + 1 for idx, name in enumerate(header) if isinstance(name, str)}

    for col_name in currency_cols:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for cell in ws[get_column_letter(col_idx)][1:]:
                if cell.value is None:
                    continue
                cell.number_format = '"$"#,##0.00;-"$"#,##0.00'

    for col_name in percent_cols:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for cell in ws[get_column_letter(col_idx)][1:]:
                if cell.value is None:
                    continue
                cell.number_format = "0.00%"


def write_workbook(
    df_union: pd.DataFrame,
    output_xlsx: str,
    include_raw: bool,
    include_monthly: bool,
    groupings: list[list[str]],
) -> Path:
    out_path = Path(output_xlsx)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    summary = build_rollup(df_union.assign(_all="All"), ["_all"]).drop(columns=["_all"])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)

        if include_monthly:
            monthly = build_rollup(df_union, ["tax_month"])
            monthly.to_excel(writer, sheet_name="By Month", index=False)

        for grouping in groupings:
            sheet_name = ("By " + ", ".join(grouping))[:31]
            rollup = build_rollup(df_union, grouping)
            rollup.to_excel(writer, sheet_name=sheet_name, index=False)

        if include_raw:
            df_union.to_excel(writer, sheet_name="Raw Union", index=False)

        wb = writer.book
        for name in wb.sheetnames:
            ws = wb[name]
            style_sheet(ws)
            if name == "Raw Union":
                apply_number_formats(
                    ws, currency_cols=["stake", "potential_payout", "bet_profit"], percent_cols=[]
                )
            else:
                apply_number_formats(
                    ws,
                    currency_cols=[
                        "handle",
                        "net_profit",
                        "gross_win_profit",
                        "gross_loss_amount",
                        "avg_bet",
                    ],
                    percent_cols=["roi"],
                )

    return out_path


# -----------------------------
# Main
# -----------------------------
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    args = parser.parse_args()

    cfg = load_config(args.config)

    # Enforce "Pikkit is required"
    pikkit_cfg = cfg.inputs.get("pikkit")
    if pikkit_cfg is None or not pikkit_cfg.enabled:
        raise SystemExit(
            "Config error: inputs.pikkit must be present and enabled=true (Pikkit is required)."
        )
    if not pikkit_cfg.csv_path:
        raise SystemExit("Config error: inputs.pikkit.csv_path is required.")

    parts: list[pd.DataFrame] = []
    stats: dict[str, dict[str, int]] = {}

    # Pikkit (required)
    pikkit_date_field = cfg.tax_year_date_field_by_source.get("pikkit", "time_settled_iso")
    df_p, st = read_pikkit(pikkit_cfg, pikkit_date_field, cfg.tax_year)
    parts.append(df_p)
    stats["pikkit"] = st

    # OddsJam (optional)
    oddsjam_cfg = cfg.inputs.get("oddsjam")
    if oddsjam_cfg is not None and oddsjam_cfg.enabled:
        df_o, st = read_oddsjam(oddsjam_cfg, cfg.tax_year)
        parts.append(df_o)
        stats["oddsjam"] = st

    # Manual (optional)
    manual_cfg = cfg.inputs.get("manual")
    if manual_cfg is not None and manual_cfg.enabled:
        df_m, st = read_manual(manual_cfg, cfg.tax_year)
        parts.append(df_m)
        stats["manual"] = st

    df_union = pd.concat(parts, ignore_index=True)

    # Final safety: ensure created_dt is Excel-safe (tz-naive) before writing
    df_union = _make_excel_safe_datetime(df_union, "created_dt")

    # Ensure stake exists for rollups
    before = len(df_union)
    df_union = df_union[df_union["stake"].notna()].copy()
    df_union["bet_profit"] = _to_float(df_union["bet_profit"])

    # Round numeric columns
    for c in ["odds", "stake", "potential_payout", "bet_profit"]:
        if c in df_union.columns:
            df_union[c] = df_union[c].round(cfg.round_decimals)

    # Render dynamic output path using tax_year template
    output_rendered = cfg.output_xlsx.format(tax_year=cfg.tax_year)

    out_path = write_workbook(
        df_union=df_union,
        output_xlsx=output_rendered,
        include_raw=cfg.include_raw_union_sheet,
        include_monthly=cfg.include_monthly_sheet,
        groupings=cfg.groupings,
    )

    print("Run stats:")
    for k, v in stats.items():
        print(f"- {k}: {v}")
    print(f"- union_rows_before_stake_filter: {before}")
    print(f"- union_rows_after_stake_filter: {len(df_union)}")
    print(f"Wrote PnL workbook: {out_path.resolve()}")


if __name__ == "__main__":
    main()
